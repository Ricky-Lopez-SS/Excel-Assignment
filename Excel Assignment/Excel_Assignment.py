#from _typeshed import NoneType
import logging
from datetime import datetime
from openpyxl.descriptors.base import String
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook, load_workbook


if __name__ == '__main__':

    #https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/ source info hehe
    #https://docs.python.org/3/howto/logging.html logging info hehehehe

    #resources/expedia_report_monthly_january_2018.xlsx
    #resources/expedia_report_monthly_march_2018.xlsx

    logging.basicConfig(filename = "resources/logfile.log", filemode='w', format='[%(asctime)s][%(levelname)s] %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', level=logging.DEBUG)

    logging.debug("this")
    logging.info("is")
    logging.warning("a")
    logging.error("test")



    filepath = input("Please enter the path of the excel document you would like to parse.\n\n")

    try:
 
        wb = load_workbook(filepath)

        months = ((1, 'jan') , (2,'feb'), (3,'mar'), (4,'apr'), (5,'may'), (6,'june'), (7,'july'), (8,'aug'), (9,'sept'), (10,'oct'), (11,'nov'), (12,'dec') )

        for m in months:
            if m[1] in filepath.lower():
                monthValue = m[0]
                monthName = m[1]
            


        ws_summary = wb['Summary Rolling MoM']
        #print(type(ws['B2'].value))

        month_row_index = 0


        #finds row of month in Summary Rolling MoM.
        for i in range(1, ws_summary.max_column + 1):

            if(isinstance(ws_summary.cell(row = i, column = 1).value , datetime)) : #is a datetime cell.

                if ws_summary.cell(row = i, column = 1).value.month == monthValue:
                    month_row_index = i
                    break


        #finds calls offered. 
        for i in range(1, ws_summary.max_row + 1) :
            if 'calls offered' in  str(ws_summary.cell(row = 1, column = i).value).lower() :
                call_offered = ws_summary.cell(row = month_row_index, column = i)

        #finds abandon after 30s.
        for i in range(1, ws_summary.max_row + 1) :
            if 'abandon after 30s' in  str(ws_summary.cell(row = 1, column = i).value).lower() :
                abandon_after_30s = ws_summary.cell(row = month_row_index, column = i)

        #finds FCR.
        for i in range(1, ws_summary.max_row + 1) :
            if 'fcr' in  str(ws_summary.cell(row = 1, column = i).value).lower() :
                fcr = ws_summary.cell(row = month_row_index, column = i)

        #finds DSAT.
        for i in range(1, ws_summary.max_row + 1) :
            if 'dsat' in  str(ws_summary.cell(row = 1, column = i).value).lower() :
                dsat = ws_summary.cell(row = month_row_index, column = i)

        #finds CSAT.
        for i in range(1, ws_summary.max_row + 1) :
            if 'csat' in  str(ws_summary.cell(row = 1, column = i).value).lower() :
                csat = ws_summary.cell(row = month_row_index, column = i)
        
        #print("{}, {}, {}, {}, {}, {}".format(month_row_index, call_offered.value, abandon_after_30s.value , fcr.value, dsat.value, csat.value) )

        print("Calls Offered: {}\n\nAbandon after 30s: {:2.2f}%\n\nFCR: {:2.2f}%\n\nDSAT: {:2.2f}%\n\nCSAT: {:2.2f}%\n\n"\
                .format(call_offered.value, abandon_after_30s.value * 100, fcr.value * 100, dsat.value * 100, csat.value * 100) )


        ws_voc = wb['VOC Rolling MoM']

        month_column_index = 0


        #Finds column of month in VOC Rolling MoM 
        for i in range(1, ws_voc.max_row + 1):
            if isinstance(ws_voc.cell(row = 1, column = i).value, datetime) : #if table value is a datetime value
                if ws_voc.cell(row = 1, column = i).value.month == monthValue:
                    month_column_index = i
                    break
            elif isinstance(ws_voc.cell(row = 1, column = i).value, str) : #if table value is a string value
                if 'mar' in ws_voc.cell(row = 1, column = i).value.lower():
                    month_column_index = i
                    break

        #harvests information from voc excel page.
        for i in range(2, ws_voc.max_column + 1):

            if(not(ws_voc.cell(row = i, column = 1).value is None)): #item is not NoneType

                if 'base size' in ws_voc.cell(row = i, column = 1).value.lower():
                    base_size = ws_voc.cell(row = i, column = month_column_index).value
                elif 'promoters' in ws_voc.cell(row = i, column = 1).value.lower():
                    promoters = ws_voc.cell(row = i, column = month_column_index).value
                elif 'passives' in ws_voc.cell(row = i, column = 1).value.lower():
                    passives = ws_voc.cell(row = i, column = month_column_index).value
                elif 'dectractors' in ws_voc.cell(row = i, column = 1).value.lower():
                    detractors = ws_voc.cell(row = i, column = month_column_index).value
                elif 'overall nps %' in ws_voc.cell(row = i, column = 1).value.lower():
                    overall_aarp = ws_voc.cell(row = i+1, column = month_column_index).value
                elif 'sat with agent %' == ws_voc.cell(row = i, column = 1).value.lower():
                    sat_with_agent_aarp = ws_voc.cell(row = i+1, column = month_column_index).value
                elif 'dsat with agent %' in ws_voc.cell(row = i, column = 1).value.lower():
                    dsat_with_agent_aarp = ws_voc.cell(row = i+1, column = month_column_index).value

        print("base size: {}\n\npromoters: {}\n\npassives: {}\n\ndetractors: {}\n\noverall nps: {}\n\nsaw with agent %: {}\n\ndsat with agent %: {}\n\n".format\
            (base_size, promoters, passives, detractors, overall_aarp, sat_with_agent_aarp, dsat_with_agent_aarp))



    except InvalidFileException:
        print("Sorry, that filepath is invalid. Please try rerunning the program with a new filepath.\n")
        exit()

    






